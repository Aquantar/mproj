from app import app, results, outputCols, predInputFormatted, originalInputFile
from flask import Flask, render_template, request, url_for, redirect, make_response
from app.functions import *
from io import BytesIO
from pickle import load, dump
import shutil
import os
import pandas as pd
import re
from flask import flash, session
from flask_httpauth import HTTPBasicAuth
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
import gzip
from datetime import datetime

app.secret_key = "dein_geheimer_schluessel"  #Für Flash-Nachrichten
auth = HTTPBasicAuth()
users = {
    "admin": generate_password_hash("meinpasswort", method='pbkdf2:sha256')
}

@auth.verify_password
def verify_password(username, password):
    if username in users and check_password_hash(users.get(username), password):
        return username
    return None

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if verify_password(username, password):
            session['user'] = username
            return redirect(url_for('monitoring'))
        else:
            flash("Falscher Benutzername oder Passwort. Bitte versuchen Sie es erneut.")
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    return render_template('logout.html')

@app.route('/', methods=['GET', 'POST'])
@app.route('/index', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.route('/prediction_results', methods=['GET', 'POST'])
def prediction_results():
    global results
    global outputCols
    global predInputFormatted
    global originalInputFile
    if request.method == 'POST':
        originalInputFile = request.files['input_prediction']
        conversionMap = load(open('models//model1//conversionMap.pkl', 'rb'))   #load conversion map of current model
        scaler = load(open('models//model1//scaler.pkl', 'rb')) #load scaler of current model
        predInputFormatted = convertPredDataToDataframe(originalInputFile) #take input file, convert into dataframe to be used for predictions
        originalInputFile = load_workbook(originalInputFile) #saving original input as workbook in this var to re-output it including predicion results later
        probaDict = {}
        for col in outputCols:  #perform prediction for each desired predicted feature
            model = load(gzip.open('models//model1//{}.pkl'.format(col), 'rb')) #load current model
            singleColResults, probaTuple = createPrediction(model, predInputFormatted, col, conversionMap, scaler) #create predictions for current feature
            results[col] = singleColResults #add predictions to results dict
            probaDict[col] = probaTuple #add probability values to probaDict
        for key, value in results.items():  #add output features to input
            predInputFormatted[key] = value

        unique_values = getUniqueValues(pd.read_excel('models//model1//currentTrainData.xlsx')) #extract all unique values for display purposes
        
        return render_template('prediction_results.html', uniqueVals=unique_values, results=results, featureCount=len(predInputFormatted.index), predictionInput=predInputFormatted.to_dict(orient='records'), probaDict=probaDict)
    return

@app.route('/prediction_results_confirmed', methods=['GET', 'POST'])
def prediction_results_confirmed():
    global results
    global predInputFormatted
    global originalInputFile
    global outputCols
    if request.method == 'POST':
        #replace predicted values with modified ones from results form
        choicesOutput = request.form
        outputDict = {}
        for key, value in choicesOutput.items(multi=True):
            valTrimmed = re.sub(r' \(\d+%\)', '', value)
            try:
                listTemp = outputDict[key]
            except:
                listTemp = []
            listTemp.append(valTrimmed)
            outputDict[key] = listTemp

        #check differences in pre and post human monitoring, and document
        totalVals = 0
        diffSelection = 0
        for key, item in outputDict.items():
            listPre = results[key]
            listPost = outputDict[key]
            totalVals = totalVals + len(listPre)
            for idx, val in enumerate(listPre):
                if listPre[idx] != listPost[idx]:
                    diffSelection+=1
        modelMetrics = pd.read_excel('models//modelData.xlsx')
        lastrow = modelMetrics.iloc[-1].tolist()
        lastrow[5] = int(lastrow[5]) + totalVals
        lastrow[6] = int(lastrow[6]) + diffSelection
        lastrow[7] = lastrow[6]/lastrow[5]
        modelMetrics.iloc[-1] = lastrow
        modelMetrics.to_excel("models//modelData.xlsx", index=False)

        #combine predictions with inputs
        predictionOutput = predInputFormatted
        predictionOutput['Prüfmittel'] = outputDict['Prüfmittel']
        predictionOutput['Stichprobenverfahren'] = outputDict['Stichprobenverfahren']
        predictionOutput['Lenkungsmethode'] = outputDict['Lenkungsmethode']

        #update stashedTrainData with new rows which don't exist in training data
        trainData = pd.read_excel('models//model1//currentTrainData.xlsx')
        trainData = trainData.astype(str)
        predictionOutput = predictionOutput.astype(str)
        newData = pd.merge(predictionOutput,trainData, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', axis=1)
        stashedData = pd.read_excel('models//stashedTrainData.xlsx')
        stashedData = stashedData.astype(str)
        stashedData = pd.concat([stashedData, newData]).drop_duplicates()
        stashedData = stashedData.replace(['nan'], [""]) 
        with pd.ExcelWriter("models//stashedTrainData.xlsx") as writer:
            stashedData.to_excel(writer, index=False)  

        #get original input file to add output vals into
        originalInputFileSheet = originalInputFile.active
        start_row = 13
        end_row = originalInputFileSheet.max_row  # or set a specific number if needed
        col_pruefmittel = 15
        col_stichprobe = 16
        col_lenkung = 18
        #map lenkungsmethode number value to text value for output
        lm_mapped= []
        mappingFrame = pd.read_excel('data//controlMethod_mappingInfo.xlsx')
        for idx, x in enumerate(outputDict['Lenkungsmethode']):
            lm_mapped.append(mappingFrame.loc[mappingFrame['Lenkungsmethode'] == int(float(outputDict['Lenkungsmethode'][idx])), 'Beschreibung'].item())
        #insert vals into input
        for i, row_num in enumerate(range(start_row, end_row + 1), start=1):
            try:
                originalInputFileSheet.cell(row=row_num, column=col_pruefmittel).value = outputDict['Prüfmittel'][i-1]
                originalInputFileSheet.cell(row=row_num, column=col_stichprobe).value = outputDict['Stichprobenverfahren'][i-1]
                originalInputFileSheet.cell(row=row_num, column=col_lenkung).value = lm_mapped[i-1]
            except:
                pass

        #initiate download of output
        sio = BytesIO()
        outputName = "output"
        originalInputFile.save(sio)
        sio.seek(0)
        resp = make_response(sio.getvalue())
        resp.headers["Content-Disposition"] = "attachment; filename={}.xlsx".format(outputName)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return resp
    return

@app.route('/model_training', methods=['GET', 'POST'])
def model_training():
    global outputCols
    if request.method == 'POST':
        try:
            dir1 = 'models//model1'
            dir2 = 'models//model2'
            dir3 = 'models//model3'

            #move existing models "down" one folder (versioning)
            for fileName in os.listdir(dir2):
                shutil.copy(os.path.join(dir2, fileName), dir3)
            for fileName in os.listdir(dir1):
                shutil.copy(os.path.join(dir1, fileName), dir2)

            #combine existing train data with new rows
            trainData = pd.read_excel('models//model1//currentTrainData.xlsx')
            trainDataNew = pd.read_excel('models//stashedTrainData.xlsx')
            trainData = pd.concat([trainData, trainDataNew]).dropna(how="all")

            #train new models
            accuracyList = []
            progress = 0
            total_steps = len(outputCols) + 2  #steps = col count + preprocessing + metrics

            #preprocessing
            progress += 1
            print(f"Preprocessing abgeschlossen ({progress}/{total_steps})")

            #model training for each output col
            for col in outputCols:
                model, scaler, conversionMap, accuracy = trainNewModel(trainData, col)
                dump(model, gzip.open(f'models//model1//{col}.pkl', 'wb'))
                dump(scaler, open('models//model1//scaler.pkl', 'wb'))
                dump(conversionMap, open('models//model1//conversionMap.pkl', 'wb'))
                accuracyList.append(accuracy)
                progress += 1
                print(f"Modell für '{col}' trainiert ({progress}/{total_steps})")

            #export training data
            trainData.to_excel("models//model1//currentTrainData.xlsx", index=False)
            pd.DataFrame(columns=list(trainData.columns.values)).to_excel("models//stashedTrainData.xlsx", index=False)

            #refresh model metrics file
            modelMetrics = pd.read_excel('models//modelData.xlsx')
            if not modelMetrics.empty and 'modelID' in modelMetrics.columns:
                last_model_id = pd.to_numeric(modelMetrics['modelID'], errors='coerce').dropna().astype(int).max()
                modelID = last_model_id + 1
            else:
                modelID = 1  #start from 1 if file is empty or column missing
            newrow = [modelID, datetime.today().strftime('%Y-%m-%d')] + accuracyList + [0, 0, 0]
            modelMetrics.loc[len(modelMetrics)] = newrow
            modelMetrics.to_excel("models//modelData.xlsx", index=False)

            progress += 1
            print(f"Modellmetriken aktualisiert ({progress}/{total_steps})")

            #prepare loading of monitoring page
            #load current model metrics
            accuracyData = [
                [float(i) for i in modelMetrics['accuracy_1'].tolist()],
                [float(i) for i in modelMetrics['accuracy_2'].tolist()],
                [float(i) for i in modelMetrics['accuracy_3'].tolist()]
            ]
            last_row = modelMetrics.iloc[-1]
            #load relevant data about current model
            last_row = modelMetrics.iloc[-1]
            predictionStats = {
                'totalPredictions': int(last_row['totalPredictions']),
                'predictionChanged': int(last_row['predictionChanged']),
                'predictionChangeRatio': f"{float(last_row['predictionChangeRatio']) * 100:.2f}%"
            }
            modelIDs = modelMetrics['modelID'].tolist()

            #output monitoring page
            return render_template("monitoring.html", accuracyData=accuracyData, modelIDs=modelIDs, predictionStats=predictionStats)

        except Exception as e:
            print(f"Fehler beim Modelltraining: {e}")
            flash(f"Fehler beim Modelltraining: {e}")
            return redirect(url_for('stasheddata'))
    return redirect(url_for('stasheddata'))

@app.route('/monitoring', methods=['GET', 'POST'])
def monitoring():
    if 'user' not in session:
        return redirect(url_for('login'))

    accuracyData = [[], [], []]
    modelIDs = []
    predictionStats = {}

    try:
        #open modelData.xlsx
        modelMetrics = pd.read_excel('models/modelData.xlsx')

        #extract relevant info for all models (graph)
        accuracyData_1 = [float(i) for i in modelMetrics['accuracy_1'].tolist()]
        accuracyData_2 = [float(i) for i in modelMetrics['accuracy_2'].tolist()]
        accuracyData_3 = [float(i) for i in modelMetrics['accuracy_3'].tolist()]
        accuracyData = [accuracyData_1, accuracyData_2, accuracyData_3]

        modelIDs = modelMetrics['modelID'].tolist()

        #extract relevant info for latest model (prediction data)
        last_row = modelMetrics.iloc[-1]
        predictionStats = {
            'totalPredictions': int(last_row['totalPredictions']),
            'predictionChanged': int(last_row['predictionChanged']),
            'predictionChangeRatio': f"{float(last_row['predictionChangeRatio']) * 100:.2f}%"
        }

    except Exception as e:
        print(f"Error when loading model data: {e}")

    return render_template("monitoring.html", accuracyData=accuracyData, modelIDs=modelIDs, predictionStats=predictionStats)

@app.route('/stasheddata', methods=['GET', 'POST'])
def stasheddata(): #load and display stashed data (new training data which has not been incorporated in a model yet)
    file_path = os.path.join("models", "stashedTrainData.xlsx")

    if not os.path.exists(file_path):
        return render_template("stasheddata.html", data=None, error="Datei nicht gefunden", row_count=0)

    try:
        df = pd.read_excel(file_path, header=0, dtype=str)
        if df.empty:
            return render_template("stasheddata.html", data=None, error="Die Datei ist leer", row_count=0)
        df = df.fillna("--")
        row_count = len(df)
        df['Aktion'] = df.index.map(
            lambda i: f'<form action="/delete_row" method="post" style="display:inline;"><input type="hidden" name="row_index" value="{i}"><button type="submit" style="padding: 10px 20px; background-color: #007BFF; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; margin: 10px;">Löschen</button></form>'
        )
        table_html = df.to_html(classes='table table-striped', index=False, escape=False)
        return render_template("stasheddata.html", data=table_html, error=None, row_count=row_count)

    except Exception as e:
        return render_template("stasheddata.html", data=None, error=str(e), row_count=0)
    
@app.route('/delete_row', methods=['POST'])
def delete_row(): #functionality to delete a row in stashed data view
    row_index = int(request.form['row_index'])
    file_path = os.path.join("models", "stashedTrainData.xlsx")

    try:
        df = pd.read_excel(file_path, header=0, dtype=str)
        df = df.drop(index=row_index)
        df.to_excel(file_path, index=False)
        return redirect(url_for('stasheddata'))
    except Exception as e:
        return render_template("stasheddata.html", data=None, error=str(e))
    
@app.route('/manage_models', methods=['GET'])
def manage_models(): #functionality to view current models
    file_path = os.path.join("models", "modelData.xlsx")

    if not os.path.exists(file_path):
        return render_template("manage_models.html", models=None, error="Datei modelData.xlsx nicht gefunden")

    try:
        df = pd.read_excel(file_path, header=0, dtype=str)
        if df.empty:
            return render_template("manage_models.html", models=None, error="Die Datei ist leer")

        #only show the last 3 rows (most recent models)
        df = df.tail(3).reset_index(drop=True)
        #convert accuracy and ratio columns to whole number percentages
        for col in ['accuracy_1', 'accuracy_2', 'accuracy_3', 'predictionChangeRatio']:
            if col in df.columns:
                df[col] = df[col].astype(float).map(lambda x: f"{round(x * 100)}%")
        #add action buttons to backup models only (not model1)
        df['Aktion'] = df['modelID'].map(lambda model_id: (
            "" if str(model_id) == "1" else
            f'''<form action="/reset_model" method="post" onsubmit="return confirm('Achtung: Diese Aktion kann nicht rückgängig gemacht werden. Fortfahren?');"><input type="hidden" name="model_id" value="{model_id}"><button type="submit" style="padding:6px 15px; background-color:#dc3545; color:white; border:none; border-radius:4px; cursor:pointer;">Wiederherstellen</button></form>'''
        ))
        df.columns = [
            "Modell-ID", "Trainingsdatum", "Genauigkeit Prüfmittel", "Genauigkeit Stichprobenverfahren",
            "Genauigkeit Lenkungsmethode", "Gesamtanzahl Vorhersagen", "Angepasste Vorhersagen",
            "Änderungsquote", "Aktion"
        ]

        #convert to html table
        models_table = df.to_html(classes='table table-striped', index=False, escape=False)
        return render_template("manage_models.html", models=models_table, error=None)

    except Exception as e:
        return render_template("manage_models.html", models=None, error=str(e))
    
@app.route('/reset_model', methods=['POST'])
def reset_model(): #logic to reset a model to a previous version
    model_id = request.form.get('model_id')
    model_name = "model" + str(model_id)
    if model_name not in ['model2', 'model3']:
        return redirect('/manage_models')  #only allow valid backups
    model_data_path = os.path.join("models", "modelData.xlsx")

    df = pd.read_excel(model_data_path, header=0, dtype=str)
    df_reversed = df[::-1].reset_index(drop=True)

    #find row of selected model
    selected_row = df_reversed[df_reversed['modelID'] == model_id]
    if selected_row.empty:
        return redirect('/manage_models')
    row_index = selected_row.index[0]

    #only keep selected model and older ones
    df_new = df_reversed.iloc[:row_index + 1][::-1]

    #mirror change to excel file
    df_new.to_excel(model_data_path, index=False)

    #reset training data to current selected model and add new "unused" rows back to stashedTrainData
    current_model_path = os.path.join("models", "model1", "currentTrainData.xlsx")
    backup_model_path = os.path.join("models", model_name, "currentTrainData.xlsx")
    stash_path = os.path.join("models", "stashedTrainData.xlsx")

    if os.path.exists(current_model_path) and os.path.exists(backup_model_path):
        current_df = pd.read_excel(current_model_path, dtype=str)
        backup_df = pd.read_excel(backup_model_path, dtype=str)

        #get all training data rows which exist in newest model, but not in the model that we have reset to
        diff_df = pd.concat([current_df, backup_df]).drop_duplicates(keep=False)

        #append rows to stashedTrainData
        if not diff_df.empty:
            if os.path.exists(stash_path):
                existing_stash_df = pd.read_excel(stash_path, dtype=str)
                updated_stash_df = pd.concat([existing_stash_df, diff_df], ignore_index=True).drop_duplicates()
            else:
                updated_stash_df = diff_df
            updated_stash_df.to_excel(stash_path, index=False)

    #replace model1 folder contents
    source_dir = os.path.join("models", model_name)
    dest_dir = os.path.join("models", "model1")
    if os.path.exists(dest_dir):
        shutil.rmtree(dest_dir)
    shutil.copytree(source_dir, dest_dir)

    #clear model2 and model3
    for backup in ['model2', 'model3']:
        backup_path = os.path.join("models", backup)
        if os.path.exists(backup_path):
            shutil.rmtree(backup_path)
            os.makedirs(backup_path)

    return redirect('/manage_models')
